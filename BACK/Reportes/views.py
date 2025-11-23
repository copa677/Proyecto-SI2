"""
Vistas para generar reportes de ventas, producción e inventario.
Compatible con exportación a PDF y Excel.
"""
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db import connection
from django.http import HttpResponse
from datetime import datetime, timedelta
from io import BytesIO
import json

# Importar librerías para PDF y Excel
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference, LineChart
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReporteVentasView(APIView):
    """
    Genera reporte de ventas (Notas de Salida) con filtros opcionales
    GET /reportes/ventas/?fecha_inicio=YYYY-MM-DD&fecha_fin=YYYY-MM-DD&formato=json|pdf|excel
    """
    def get(self, request):
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        formato = request.query_params.get('formato', 'json')
        
        try:
            # Consulta SQL para obtener datos de ventas
            with connection.cursor() as cursor:
                query = """
                    SELECT 
                        ns.id_salida,
                        ns.fecha_salida,
                        COALESCE(p.nombre_completo, 'Sin Asignar') as responsable,
                        dns.nombre_materia_prima as producto,
                        COALESCE(l.codigo_lote, 'Sin Lote') as lote_asociado,
                        dns.cantidad,
                        dns.unidad_medida,
                        ns.motivo,
                        ns.estado,
                        0.00 as precio_total
                    FROM nota_salida ns
                    INNER JOIN detalle_nota_salida dns ON ns.id_salida = dns.id_salida
                    LEFT JOIN lotes l ON dns.id_lote = l.id_lote
                    LEFT JOIN personal p ON ns.id_personal = p.id
                    WHERE 1=1
                """
                
                params = []
                if fecha_inicio:
                    query += " AND ns.fecha_salida >= %s"
                    params.append(fecha_inicio)
                if fecha_fin:
                    query += " AND ns.fecha_salida <= %s"
                    params.append(fecha_fin)
                
                query += " ORDER BY ns.fecha_salida DESC, ns.id_salida"
                
                cursor.execute(query, params)
                columns = [col[0] for col in cursor.description]
                ventas = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Convertir Decimal a float para JSON
            for venta in ventas:
                if 'cantidad' in venta and venta['cantidad'] is not None:
                    venta['cantidad'] = float(venta['cantidad'])
                if 'precio_total' in venta and venta['precio_total'] is not None:
                    venta['precio_total'] = float(venta['precio_total'])
                if 'fecha_salida' in venta and venta['fecha_salida'] is not None:
                    venta['fecha_salida'] = venta['fecha_salida'].isoformat()
            
            # Calcular estadísticas avanzadas
            total_ventas = len(ventas)
            cantidad_total = sum(v.get('cantidad', 0) for v in ventas)
            monto_total = sum(v.get('precio_total', 0) for v in ventas)
            
            # Agrupar por producto
            productos = {}
            for v in ventas:
                prod = v.get('producto', 'Desconocido')
                if prod not in productos:
                    productos[prod] = {'cantidad': 0, 'ventas': 0}
                productos[prod]['cantidad'] += v.get('cantidad', 0)
                productos[prod]['ventas'] += 1
            
            # Top 5 productos
            top_productos = sorted(productos.items(), key=lambda x: x[1]['cantidad'], reverse=True)[:5]
            
            # Agrupar por responsable
            responsables = {}
            for v in ventas:
                resp = v.get('responsable', 'Sin Asignar')
                if resp not in responsables:
                    responsables[resp] = 0
                responsables[resp] += 1
            
            # Agrupar por estado
            estados = {}
            for v in ventas:
                est = v.get('estado', 'Desconocido')
                if est not in estados:
                    estados[est] = 0
                estados[est] += 1
            
            # Promedio de ventas
            promedio_cantidad = cantidad_total / total_ventas if total_ventas > 0 else 0
            
            data = {
                'ventas': ventas,
                'estadisticas': {
                    'total_ventas': total_ventas,
                    'cantidad_total': cantidad_total,
                    'monto_total': monto_total,
                    'promedio_cantidad': promedio_cantidad,
                    'top_productos': top_productos,
                    'por_responsable': responsables,
                    'por_estado': estados
                },
                'filtros': {
                    'fecha_inicio': fecha_inicio,
                    'fecha_fin': fecha_fin
                },
                'fecha_generacion': datetime.now().isoformat()
            }
            
            # Exportar según formato
            if formato == 'pdf':
                return self.exportar_pdf(data)
            elif formato == 'excel':
                return self.exportar_excel(data)
            else:
                return Response(data, status=status.HTTP_200_OK)
                
        except Exception as e:
            return Response(
                {'error': f'Error al generar reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def exportar_pdf(self, data):
        """Exporta el reporte a PDF con gráficas"""
        if not REPORTLAB_AVAILABLE:
            return Response(
                {'error': 'ReportLab no está instalado'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # =======ENCABEZADO=======
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#4F46E5'),
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#6B7280'),
            alignment=TA_CENTER,
            spaceAfter=12
        )
        
        elements.append(Paragraph('📊 REPORTE DE VENTAS', title_style))
        elements.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}", subtitle_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # =======TARJETAS DE ESTADÍSTICAS=======
        stats = data['estadisticas']
        stats_data = [
            ['INDICADORES CLAVE', '', '', ''],
            ['Total Ventas', 'Cantidad Total', 'Monto Total', 'Promedio por Venta'],
            [
                f"{stats['total_ventas']}",
                f"{stats['cantidad_total']:.2f}",
                f"Bs. {stats['monto_total']:.2f}",
                f"{stats['promedio_cantidad']:.2f}"
            ]
        ]
        
        stats_table = Table(stats_data, colWidths=[1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#E0E7FF')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 9),
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#F3F4F6')),
            ('FONTSIZE', (0, 2), (-1, 2), 12),
            ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 2), (-1, 2), colors.HexColor('#1F2937')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#D1D5DB')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12)
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # =======GRÁFICA TOP 5 PRODUCTOS=======
        if stats.get('top_productos'):
            elements.append(Paragraph('<b>📈 TOP 5 PRODUCTOS MÁS VENDIDOS</b>', styles['Heading2']))
            elements.append(Spacer(1, 10))
            
            # Crear gráfica de barras
            drawing = Drawing(500, 200)
            bc = VerticalBarChart()
            bc.x = 50
            bc.y = 50
            bc.height = 125
            bc.width = 400
            
            top_5 = stats['top_productos'][:5]
            bc.data = [[item[1]['cantidad'] for item in top_5]]
            bc.categoryAxis.categoryNames = [item[0][:15] for item in top_5]
            bc.categoryAxis.labels.fontSize = 8
            bc.categoryAxis.labels.angle = 45   
            bc.valueAxis.valueMin = 0
            bc.bars[0].fillColor = colors.HexColor('#4F46E5')
            
            drawing.add(bc)
            elements.append(drawing)
            elements.append(Spacer(1, 0.2*inch))
        
        # =======DISTRIBUCIÓN POR ESTADO=======
        if stats.get('por_estado'):
            elements.append(Paragraph('<b>📊 DISTRIBUCIÓN POR ESTADO</b>', styles['Heading2']))
            elements.append(Spacer(1, 10))
            
            estado_data = [['Estado', 'Cantidad', 'Porcentaje']]
            total = sum(stats['por_estado'].values())
            for estado, cant in stats['por_estado'].items():
                porcentaje = (cant / total * 100) if total > 0 else 0
                estado_data.append([estado, str(cant), f"{porcentaje:.1f}%"])
            
            estado_table = Table(estado_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
            estado_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#D1D5DB')),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB'))
            ]))
            elements.append(estado_table)
            elements.append(PageBreak())
        
        # =======TABLA DETALLADA DE VENTAS=======
        elements.append(Paragraph('<b>📋 DETALLE DE VENTAS</b>', styles['Heading2']))
        elements.append(Spacer(1, 10))
        
        table_data = [['ID', 'Fecha', 'Producto', 'Responsable', 'Cantidad', 'Estado']]
        for venta in data['ventas'][:50]:  # Limitar a 50 registros
            table_data.append([
                str(venta.get('id_salida', ''))[:8],
                venta.get('fecha_salida', '')[:10] if venta.get('fecha_salida') else '',
                venta.get('producto', '')[:25],
                venta.get('responsable', '')[:20],
                f"{venta.get('cantidad', 0):.2f}",
                venta.get('estado', '')[:15]
            ])
        
        detail_table = Table(table_data, colWidths=[0.6*inch, 1*inch, 2*inch, 1.5*inch, 0.8*inch, 1*inch])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#F9FAFB'), colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB'))
        ]))
        elements.append(detail_table)
        
        # =======PIE DE PÁGINA=======
        elements.append(Spacer(1, 0.3*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#9CA3AF'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"Reporte generado por ManufacturaPRO - {datetime.now().year}", footer_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
    
    def exportar_excel(self, data):
        """Exporta el reporte a Excel con formato profesional"""
        if not OPENPYXL_AVAILABLE:
            return Response(
                {'error': 'openpyxl no está instalado. Instala con: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        wb = Workbook()
        
        # === HOJA 1: RESUMEN ===
        ws_resumen = wb.active
        ws_resumen.title = "📊 Resumen"
        
        # Título principal con fondo de color
        ws_resumen['A1'] = '💰 REPORTE DE VENTAS'
        ws_resumen['A1'].font = Font(bold=True, size=24, color="FFFFFF")
        ws_resumen['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_resumen['A1'].fill = PatternFill(start_color="4F46E5", end_color="7C3AED", fill_type="solid")
        ws_resumen.merge_cells('A1:F1')
        ws_resumen.row_dimensions[1].height = 40
        
        # Agregar borde grueso al título
        thick_border = Border(
            left=Side(style='thick', color='4F46E5'),
            right=Side(style='thick', color='4F46E5'),
            top=Side(style='thick', color='4F46E5'),
            bottom=Side(style='thick', color='4F46E5')
        )
        ws_resumen['A1'].border = thick_border
        
        # Fecha de generación con fondo
        ws_resumen['A2'] = f'📅 Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws_resumen['A2'].font = Font(italic=True, size=11, color="FFFFFF", bold=True)
        ws_resumen['A2'].fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        ws_resumen['A2'].alignment = Alignment(horizontal='center')
        ws_resumen.merge_cells('A2:F2')
        ws_resumen.row_dimensions[2].height = 25
        
        # Estadísticas con formato
        stats = data['estadisticas']
        row = 4
        
        # KPIs principales con colores vibrantes y iconos
        kpis = [
            ('📊 Total Ventas', stats['total_ventas'], '5B21B6', 'DDD6FE', 'FFFFFF'),
            ('📦 Cantidad Total Vendida', stats['cantidad_total'], '059669', 'A7F3D0', 'FFFFFF'),
            ('💵 Monto Total (Bs.)', f"{stats['monto_total']:.2f}", 'D97706', 'FDE68A', 'FFFFFF'),
            ('📈 Promedio por Venta', f"{stats.get('promedio_cantidad', 0):.2f}", 'DC2626', 'FECACA', 'FFFFFF')
        ]
        
        for label, value, label_color, value_color, text_color in kpis:
            # Celda de etiqueta
            ws_resumen[f'A{row}'] = label
            ws_resumen[f'A{row}'].font = Font(bold=True, size=13, color=text_color)
            ws_resumen[f'A{row}'].fill = PatternFill(start_color=label_color, end_color=label_color, fill_type="solid")
            ws_resumen[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws_resumen[f'A{row}'].border = Border(
                left=Side(style='thick', color=label_color),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thick', color=label_color),
                bottom=Side(style='thick', color=label_color)
            )
            
            # Celda de valor
            ws_resumen[f'B{row}'] = value
            ws_resumen[f'B{row}'].font = Font(size=16, bold=True, color='1E293B')
            ws_resumen[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws_resumen[f'B{row}'].fill = PatternFill(start_color=value_color, end_color=value_color, fill_type="solid")
            ws_resumen[f'B{row}'].border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thick', color=label_color),
                top=Side(style='thick', color=label_color),
                bottom=Side(style='thick', color=label_color)
            )
            
            ws_resumen.row_dimensions[row].height = 30
            row += 1
        
        # Top 5 Productos
        if stats.get('top_productos'):
            row += 2
            ws_resumen[f'A{row}'] = '🏆 TOP 5 PRODUCTOS MÁS VENDIDOS'
            ws_resumen[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
            ws_resumen[f'A{row}'].fill = PatternFill(start_color="DC2626", end_color="EF4444", fill_type="solid")
            ws_resumen[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws_resumen.merge_cells(f'A{row}:D{row}')
            ws_resumen.row_dimensions[row].height = 30
            row += 1
            
            headers = ['🎯 Producto', '📦 Cantidad', '💰 Ventas', '📊 %']
            header_colors = ['FB923C', 'FBBF24', '34D399', '60A5FA']
            for col_num, (header, color) in enumerate(zip(headers, header_colors), 1):
                cell = ws_resumen.cell(row=row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='medium', color='1F2937'),
                    right=Side(style='medium', color='1F2937'),
                    top=Side(style='medium', color='1F2937'),
                    bottom=Side(style='medium', color='1F2937')
                )
            ws_resumen.row_dimensions[row].height = 25
            row += 1
            
            # Colores degradados para las filas
            row_colors = ['FEF3C7', 'FDE68A', 'FCD34D', 'FBBF24', 'F59E0B']
            for idx, (prod, info) in enumerate(stats['top_productos']):
                porcentaje = (info['cantidad'] / stats['cantidad_total'] * 100) if stats['cantidad_total'] > 0 else 0
                bg_color = row_colors[idx] if idx < len(row_colors) else 'FFFFFF'
                
                # Emoji de medalla para top 3
                medal = ['🥇', '🥈', '🥉'][idx] if idx < 3 else '⭐'
                
                ws_resumen[f'A{row}'] = f"{medal} {prod[:28]}"
                ws_resumen[f'B{row}'] = info['cantidad']
                ws_resumen[f'C{row}'] = info['ventas']
                ws_resumen[f'D{row}'] = f"{porcentaje:.1f}%"
                
                for col in ['A', 'B', 'C', 'D']:
                    cell = ws_resumen[f'{col}{row}']
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    cell.font = Font(bold=True if idx < 3 else False, size=11)
                    cell.alignment = Alignment(horizontal='left' if col == 'A' else 'center', vertical='center')
                    cell.border = Border(
                        left=Side(style='medium', color='94A3B8'),
                        right=Side(style='medium', color='94A3B8'),
                        top=Side(style='thin', color='CBD5E1'),
                        bottom=Side(style='thin', color='CBD5E1')
                    )
                ws_resumen.row_dimensions[row].height = 22
                row += 1
        
        # Ajustar anchos
        ws_resumen.column_dimensions['A'].width = 42
        ws_resumen.column_dimensions['B'].width = 20
        ws_resumen.column_dimensions['C'].width = 18
        ws_resumen.column_dimensions['D'].width = 15
        
        # === HOJA 2: DETALLE DE VENTAS ===
        ws_detalle = wb.create_sheet("📋 Detalle Ventas")
        
        # Título de la hoja
        ws_detalle['A1'] = '📋 DETALLE COMPLETO DE VENTAS'
        ws_detalle['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws_detalle['A1'].fill = PatternFill(start_color="1E40AF", end_color="3B82F6", fill_type="solid")
        ws_detalle['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_detalle.merge_cells('A1:H1')
        ws_detalle.row_dimensions[1].height = 35
        
        # Encabezados con iconos
        headers = ['🆔 ID', '📅 Fecha', '📦 Producto', '🏷️ Lote', '📊 Cantidad', '💰 Total (Bs.)', '👤 Responsable', '✅ Estado']
        header_colors = ['6366F1', '8B5CF6', 'EC4899', 'F43F5E', 'F59E0B', '10B981', '06B6D4', '14B8A6']
        
        for col_num, (header, color) in enumerate(zip(headers, header_colors), 1):
            cell = ws_detalle.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thick', color='1F2937'),
                right=Side(style='thick', color='1F2937'),
                top=Side(style='thick', color='1F2937'),
                bottom=Side(style='thick', color='1F2937')
            )
        ws_detalle.row_dimensions[2].height = 30
        
        # Datos con formato alternado y colores vibrantes
        for row_num, venta in enumerate(data['ventas'], 3):
            ws_detalle.cell(row=row_num, column=1, value=venta['id_salida'])
            ws_detalle.cell(row=row_num, column=2, value=venta['fecha_salida'][:10] if venta['fecha_salida'] else '')
            ws_detalle.cell(row=row_num, column=3, value=venta['producto'])
            ws_detalle.cell(row=row_num, column=4, value=venta['lote_asociado'])
            ws_detalle.cell(row=row_num, column=5, value=venta['cantidad'])
            ws_detalle.cell(row=row_num, column=6, value=venta['precio_total'])
            ws_detalle.cell(row=row_num, column=7, value=venta.get('responsable', 'N/A'))
            ws_detalle.cell(row=row_num, column=8, value=venta.get('estado', 'Completado'))
            
            # Formato de moneda
            ws_detalle.cell(row=row_num, column=6).number_format = 'Bs. #,##0.00'
            ws_detalle.cell(row=row_num, column=6).font = Font(bold=True, color="059669", size=11)
            
            # Colores alternados más vibrantes
            if row_num % 2 == 0:
                fill_color = 'EFF6FF'  # Azul muy claro
            else:
                fill_color = 'F0FDF4'  # Verde muy claro
            
            for col in range(1, 9):
                cell = ws_detalle.cell(row=row_num, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(horizontal='center' if col != 3 else 'left', vertical='center')
                cell.font = Font(size=10)
                cell.border = Border(
                    left=Side(style='medium', color='94A3B8'),
                    right=Side(style='medium', color='94A3B8'),
                    top=Side(style='thin', color='CBD5E1'),
                    bottom=Side(style='thin', color='CBD5E1')
                )
            ws_detalle.row_dimensions[row_num].height = 22
        
        # Ajustar anchos
        ws_detalle.column_dimensions['A'].width = 10
        ws_detalle.column_dimensions['B'].width = 14
        ws_detalle.column_dimensions['C'].width = 35
        ws_detalle.column_dimensions['D'].width = 18
        ws_detalle.column_dimensions['E'].width = 14
        ws_detalle.column_dimensions['F'].width = 18
        ws_detalle.column_dimensions['G'].width = 28
        ws_detalle.column_dimensions['H'].width = 16
        
        # Congelar paneles (primera fila y segunda fila)
        ws_detalle.freeze_panes = 'A3'
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Reporte_Ventas_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response


class ReporteProduccionView(APIView):
    """
    Genera reporte de producción (Órdenes de Producción)
    GET /reportes/produccion/?fecha_inicio=YYYY-MM-DD&fecha_fin=YYYY-MM-DD&formato=json|pdf|excel
    """
    def get(self, request):
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        formato = request.query_params.get('formato', 'json')
        
        try:
            with connection.cursor() as cursor:
                query = """
                    SELECT 
                        op.id_orden,
                        op.cod_orden,
                        op.fecha_inicio,
                        op.fecha_fin,
                        op.fecha_entrega,
                        op.estado,
                        op.producto_modelo,
                        op.color,
                        op.talla,
                        op.cantidad_total,
                        COALESCE(p.nombre_completo, 'Sin Asignar') as responsable,
                        CASE 
                            WHEN op.fecha_fin < CURRENT_DATE AND op.estado != 'Completada' THEN 'Retrasada'
                            WHEN op.fecha_fin >= CURRENT_DATE THEN 'En Tiempo'
                            ELSE 'Completada'
                        END as cumplimiento
                    FROM orden_produccion op
                    LEFT JOIN personal p ON op.id_personal = p.id
                    WHERE 1=1
                """
                
                params = []
                if fecha_inicio:
                    query += " AND op.fecha_inicio >= %s"
                    params.append(fecha_inicio)
                if fecha_fin:
                    query += " AND op.fecha_fin <= %s"
                    params.append(fecha_fin)
                
                query += " ORDER BY op.fecha_inicio DESC"
                
                cursor.execute(query, params)
                columns = [col[0] for col in cursor.description]
                ordenes = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Convertir fechas a string para JSON
            for orden in ordenes:
                for campo in ['fecha_inicio', 'fecha_fin', 'fecha_entrega']:
                    if campo in orden and orden[campo] is not None:
                        orden[campo] = orden[campo].isoformat()
            
            # Estadísticas avanzadas
            total_ordenes = len(ordenes)
            completadas = sum(1 for o in ordenes if o['estado'] == 'Completada')
            en_proceso = sum(1 for o in ordenes if o['estado'] == 'En Proceso')
            retrasadas = sum(1 for o in ordenes if o['cumplimiento'] == 'Retrasada')
            cantidad_total = sum(o.get('cantidad_total', 0) for o in ordenes)
            
            # Agrupar por producto
            por_producto = {}
            for o in ordenes:
                prod = o.get('producto_modelo', 'Desconocido')
                if prod not in por_producto:
                    por_producto[prod] = {'cantidad': 0, 'ordenes': 0}
                por_producto[prod]['cantidad'] += o.get('cantidad_total', 0)
                por_producto[prod]['ordenes'] += 1
            
            # Agrupar por estado
            por_estado = {}
            for o in ordenes:
                est = o.get('estado', 'Desconocido')
                if est not in por_estado:
                    por_estado[est] = 0
                por_estado[est] += 1
            
            # Tasa de cumplimiento
            tasa_completadas = (completadas / total_ordenes * 100) if total_ordenes > 0 else 0
            
            data = {
                'ordenes': ordenes,
                'estadisticas': {
                    'total_ordenes': total_ordenes,
                    'completadas': completadas,
                    'en_proceso': en_proceso,
                    'retrasadas': retrasadas,
                    'cantidad_total_producida': cantidad_total,
                    'tasa_completadas': tasa_completadas,
                    'por_producto': por_producto,
                    'por_estado': por_estado
                },
                'filtros': {
                    'fecha_inicio': fecha_inicio,
                    'fecha_fin': fecha_fin
                },
                'fecha_generacion': datetime.now().isoformat()
            }
            
            if formato == 'pdf':
                return self.exportar_pdf(data)
            elif formato == 'excel':
                return self.exportar_excel(data)
            else:
                return Response(data, status=status.HTTP_200_OK)
                
        except Exception as e:
            return Response(
                {'error': f'Error al generar reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def exportar_pdf(self, data):
        """Exporta el reporte a PDF con gráficas"""
        if not REPORTLAB_AVAILABLE:
            return Response({'error': 'ReportLab no disponible'}, status=500)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=50, rightMargin=50, topMargin=50, bottomMargin=50)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título principal
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=22, textColor=colors.HexColor('#4F46E5'), spaceAfter=30, alignment=TA_CENTER, fontName='Helvetica-Bold')
        elements.append(Paragraph('📊 Reporte de Producción', title_style))
        elements.append(Spacer(1, 12))
        
        # Fecha de generación
        fecha_style = ParagraphStyle('Fecha', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", fecha_style))
        elements.append(Spacer(1, 20))
        
        stats = data['estadisticas']
        
        # Tarjetas de estadísticas (4 KPIs)
        stats_data = [
            [
                Paragraph(f"<para align=center><b><font size=24 color='#4F46E5'>{stats['total_ordenes']}</font></b><br/><font size=10 color='gray'>Total Órdenes</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=24 color='#10B981'>{stats['completadas']}</font></b><br/><font size=10 color='gray'>Completadas</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=24 color='#F59E0B'>{stats['en_proceso']}</font></b><br/><font size=10 color='gray'>En Proceso</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=24 color='#EF4444'>{stats['retrasadas']}</font></b><br/><font size=10 color='gray'>Retrasadas</font></para>", styles['Normal'])
            ]
        ]
        stats_table = Table(stats_data, colWidths=[1.5*inch]*4)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#EEF2FF')),
            ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#D1FAE5')),
            ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#FEF3C7')),
            ('BACKGROUND', (3, 0), (3, 0), colors.HexColor('#FEE2E2')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROUNDEDCORNERS', [10, 10, 10, 10]),
            ('LEFTPADDING', (0, 0), (-1, -1), 15),
            ('RIGHTPADDING', (0, 0), (-1, -1), 15),
            ('TOPPADDING', (0, 0), (-1, -1), 15),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 30))
        
        # Gráfica de barras por estado
        if stats.get('por_estado'):
            subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#4F46E5'), spaceAfter=10)
            elements.append(Paragraph('📈 Distribución por Estado', subtitle_style))
            elements.append(Spacer(1, 10))
            
            drawing = Drawing(400, 200)
            bc = VerticalBarChart()
            bc.x = 50
            bc.y = 20
            bc.height = 150
            bc.width = 300
            bc.data = [list(stats['por_estado'].values())]
            bc.categoryAxis.categoryNames = list(stats['por_estado'].keys())
            bc.bars[0].fillColor = colors.HexColor('#4F46E5')
            bc.valueAxis.valueMin = 0
            bc.categoryAxis.labels.boxAnchor = 'n'
            bc.categoryAxis.labels.angle = 30
            bc.valueAxis.labels.fontName = 'Helvetica'
            drawing.add(bc)
            elements.append(drawing)
            elements.append(Spacer(1, 20))
        
        # Tabla de productos
        if stats.get('por_producto'):
            elements.append(Paragraph('🏭 Top Productos', subtitle_style))
            elements.append(Spacer(1, 10))
            
            prod_data = [['Producto', 'Órdenes', 'Cantidad']]
            sorted_prods = sorted(stats['por_producto'].items(), key=lambda x: x[1]['cantidad'], reverse=True)[:5]
            for prod, info in sorted_prods:
                prod_data.append([prod[:30], str(info['ordenes']), str(info['cantidad'])])
            
            prod_table = Table(prod_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
            prod_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')])
            ]))
            elements.append(prod_table)
            elements.append(Spacer(1, 25))
        
        # Tabla detallada de órdenes
        elements.append(Paragraph('📋 Detalle de Órdenes de Producción', subtitle_style))
        elements.append(Spacer(1, 10))
        
        table_data = [['Código', 'Producto', 'Cantidad', 'Estado', 'F. Inicio', 'F. Fin']]
        for orden in data['ordenes']:
            table_data.append([
                orden['cod_orden'][:12],
                orden['producto_modelo'][:20],
                str(orden['cantidad_total']),
                orden['estado'][:15],
                orden['fecha_inicio'][:10],
                orden['fecha_fin'][:10]
            ])
        
        table = Table(table_data, colWidths=[1*inch, 1.8*inch, 0.9*inch, 1*inch, 0.9*inch, 0.9*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(f"TextilTech © {datetime.now().year} - Reporte generado automáticamente", footer_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Produccion_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response
    
    def exportar_excel(self, data):
        """Exporta el reporte a Excel con formato profesional"""
        if not OPENPYXL_AVAILABLE:
            return Response({'error': 'openpyxl no disponible'}, status=500)
        
        wb = Workbook()
        
        # === HOJA 1: RESUMEN ===
        ws_resumen = wb.active
        ws_resumen.title = "📊 Resumen"
        
        # Título con gradiente
        ws_resumen['A1'] = '🏭 REPORTE DE PRODUCCIÓN'
        ws_resumen['A1'].font = Font(bold=True, size=24, color="FFFFFF")
        ws_resumen['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_resumen['A1'].fill = PatternFill(start_color="7C3AED", end_color="A78BFA", fill_type="solid")
        ws_resumen.merge_cells('A1:F1')
        ws_resumen.row_dimensions[1].height = 40
        thick_border = Border(
            left=Side(style='thick', color='7C3AED'),
            right=Side(style='thick', color='7C3AED'),
            top=Side(style='thick', color='7C3AED'),
            bottom=Side(style='thick', color='7C3AED')
        )
        ws_resumen['A1'].border = thick_border
        
        # Fecha
        ws_resumen['A2'] = f'📅 Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws_resumen['A2'].font = Font(italic=True, size=11, color="FFFFFF", bold=True)
        ws_resumen['A2'].fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        ws_resumen['A2'].alignment = Alignment(horizontal='center')
        ws_resumen.merge_cells('A2:F2')
        ws_resumen.row_dimensions[2].height = 25
        
        # Estadísticas
        stats = data['estadisticas']
        row = 4
        
        kpis = [
            ('📊 Total Órdenes', stats['total_ordenes'], '7C3AED', 'DDD6FE', 'FFFFFF'),
            ('✅ Completadas', stats['completadas'], '059669', 'A7F3D0', 'FFFFFF'),
            ('⏳ En Proceso', stats['en_proceso'], 'D97706', 'FDE68A', 'FFFFFF'),
            ('⚠️ Retrasadas', stats['retrasadas'], 'DC2626', 'FECACA', 'FFFFFF'),
            ('📈 Tasa Completadas', f"{stats.get('tasa_completadas', 0):.1f}%", '2563EB', 'DBEAFE', 'FFFFFF')
        ]
        
        for label, value, label_color, value_color, text_color in kpis:
            ws_resumen[f'A{row}'] = label
            ws_resumen[f'A{row}'].font = Font(bold=True, size=13, color=text_color)
            ws_resumen[f'A{row}'].fill = PatternFill(start_color=label_color, end_color=label_color, fill_type="solid")
            ws_resumen[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws_resumen[f'A{row}'].border = Border(
                left=Side(style='thick', color=label_color),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thick', color=label_color),
                bottom=Side(style='thick', color=label_color)
            )
            
            ws_resumen[f'B{row}'] = value
            ws_resumen[f'B{row}'].font = Font(size=16, bold=True, color='1E293B')
            ws_resumen[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws_resumen[f'B{row}'].fill = PatternFill(start_color=value_color, end_color=value_color, fill_type="solid")
            ws_resumen[f'B{row}'].border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thick', color=label_color),
                top=Side(style='thick', color=label_color),
                bottom=Side(style='thick', color=label_color)
            )
            
            ws_resumen.row_dimensions[row].height = 30
            row += 1
        
        # Top productos
        if stats.get('por_producto'):
            row += 2
            ws_resumen[f'A{row}'] = 'PRODUCCIÓN POR PRODUCTO'
            ws_resumen[f'A{row}'].font = Font(bold=True, size=12, color="4F46E5")
            ws_resumen.merge_cells(f'A{row}:D{row}')
            row += 1
            
            headers = ['Producto', 'Cantidad', 'Órdenes']
            for col_num, header in enumerate(headers, 1):
                cell = ws_resumen.cell(row=row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            for prod, info in sorted(stats['por_producto'].items(), key=lambda x: x[1]['cantidad'], reverse=True)[:10]:
                ws_resumen[f'A{row}'] = prod[:35]
                ws_resumen[f'B{row}'] = info['cantidad']
                ws_resumen[f'C{row}'] = info['ordenes']
                row += 1
        
        ws_resumen.column_dimensions['A'].width = 40
        ws_resumen.column_dimensions['B'].width = 15
        ws_resumen.column_dimensions['C'].width = 15
        
        # === HOJA 2: DETALLE DE ÓRDENES ===
        ws_detalle = wb.create_sheet("Detalle de Órdenes")
        
        headers = ['Código', 'Producto', 'Color', 'Talla', 'Cantidad', 'Estado', 'F. Inicio', 'F. Fin', 'Responsable']
        for col_num, header in enumerate(headers, 1):
            cell = ws_detalle.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        for row_num, orden in enumerate(data['ordenes'], 2):
            ws_detalle.cell(row=row_num, column=1, value=orden['cod_orden'])
            ws_detalle.cell(row=row_num, column=2, value=orden['producto_modelo'])
            ws_detalle.cell(row=row_num, column=3, value=orden.get('color', ''))
            ws_detalle.cell(row=row_num, column=4, value=orden.get('talla', ''))
            ws_detalle.cell(row=row_num, column=5, value=orden['cantidad_total'])
            ws_detalle.cell(row=row_num, column=6, value=orden['estado'])
            ws_detalle.cell(row=row_num, column=7, value=orden['fecha_inicio'][:10] if orden['fecha_inicio'] else '')
            ws_detalle.cell(row=row_num, column=8, value=orden['fecha_fin'][:10] if orden['fecha_fin'] else '')
            ws_detalle.cell(row=row_num, column=9, value=orden.get('responsable', 'N/A'))
            
            # Color alternado
            fill_color = 'F5F5F5' if row_num % 2 == 0 else 'FFFFFF'
            for col in range(1, 10):
                cell = ws_detalle.cell(row=row_num, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
        
        ws_detalle.column_dimensions['A'].width = 15
        ws_detalle.column_dimensions['B'].width = 30
        ws_detalle.column_dimensions['C'].width = 12
        ws_detalle.column_dimensions['D'].width = 10
        ws_detalle.column_dimensions['E'].width = 12
        ws_detalle.column_dimensions['F'].width = 15
        ws_detalle.column_dimensions['G'].width = 12
        ws_detalle.column_dimensions['H'].width = 12
        ws_detalle.column_dimensions['I'].width = 25
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Produccion_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response


class ReporteInventarioView(APIView):
    """
    Genera reporte de inventario y consumo de materiales
    GET /reportes/inventario-consumo/?fecha_inicio=YYYY-MM-DD&fecha_fin=YYYY-MM-DD&formato=json|pdf|excel
    """
    def get(self, request):
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        formato = request.query_params.get('formato', 'json')
        
        try:
            # Stock de materias primas
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT 
                        i.nombre_materia_prima,
                        SUM(i.cantidad_actual) as stock_total,
                        i.unidad_medida,
                        i.estado,
                        SUM(i.stock_minimo) as stock_minimo
                    FROM inventario i
                    GROUP BY i.nombre_materia_prima, i.unidad_medida, i.estado
                    ORDER BY i.nombre_materia_prima
                """)
                columns = [col[0] for col in cursor.description]
                stock_materias = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Convertir Decimal a float
            for item in stock_materias:
                if 'stock_total' in item and item['stock_total'] is not None:
                    item['stock_total'] = float(item['stock_total'])
                if 'stock_minimo' in item and item['stock_minimo'] is not None:
                    item['stock_minimo'] = float(item['stock_minimo'])
            
            # Stock de productos terminados (órdenes completadas)
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT 
                        op.producto_modelo,
                        op.color,
                        op.talla,
                        SUM(op.cantidad_total) as cantidad
                    FROM orden_produccion op
                    WHERE op.estado = 'Completada'
                    GROUP BY op.producto_modelo, op.color, op.talla
                    ORDER BY op.producto_modelo
                """)
                columns = [col[0] for col in cursor.description]
                stock_productos = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Consumo de materiales (notas de salida)
            with connection.cursor() as cursor:
                query = """
                    SELECT 
                        dns.nombre_materia_prima as nombre_materia,
                        SUM(dns.cantidad) as consumo_total,
                        dns.unidad_medida
                    FROM detalle_nota_salida dns
                    INNER JOIN nota_salida ns ON dns.id_salida = ns.id_salida
                    WHERE 1=1
                """
                
                params = []
                if fecha_inicio:
                    query += " AND ns.fecha_salida >= %s"
                    params.append(fecha_inicio)
                if fecha_fin:
                    query += " AND ns.fecha_salida <= %s"
                    params.append(fecha_fin)
                
                query += " GROUP BY dns.nombre_materia_prima, dns.unidad_medida ORDER BY consumo_total DESC"
                
                cursor.execute(query, params)
                columns = [col[0] for col in cursor.description]
                consumo_materiales = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Convertir Decimal a float
            for item in consumo_materiales:
                if 'consumo_total' in item and item['consumo_total'] is not None:
                    item['consumo_total'] = float(item['consumo_total'])
            
            # Estadísticas avanzadas
            total_materias = len(stock_materias)
            stock_bajo = sum(1 for m in stock_materias if m.get('stock_total', 0) < m.get('stock_minimo', 0))
            stock_disponible = sum(1 for m in stock_materias if m.get('estado') == 'Disponible')
            consumo_total = sum(c.get('consumo_total', 0) for c in consumo_materiales)
            
            # Agrupar por estado
            por_estado = {}
            for m in stock_materias:
                est = m.get('estado', 'Desconocido')
                if est not in por_estado:
                    por_estado[est] = 0
                por_estado[est] += 1
            
            # Top materiales con stock bajo
            materiales_criticos = [
                m for m in stock_materias 
                if m.get('stock_total', 0) < m.get('stock_minimo', 0)
            ][:5]
            
            data = {
                'fecha_generacion': datetime.now().isoformat(),
                'stock_materias_primas': stock_materias,
                'stock_productos_terminados': [
                    {
                        'producto': f"{p['producto_modelo']} - {p['color']} - {p['talla']}",
                        'cantidad': p['cantidad']
                    }
                    for p in stock_productos
                ],
                'consumo_materiales': {
                    'filtros': {
                        'fecha_inicio': fecha_inicio,
                        'fecha_fin': fecha_fin
                    },
                    'data': consumo_materiales
                },
                'estadisticas': {
                    'total_materias': total_materias,
                    'stock_bajo': stock_bajo,
                    'stock_disponible': stock_disponible,
                    'consumo_total': consumo_total,
                    'por_estado': por_estado,
                    'materiales_criticos': materiales_criticos
                }
            }
            
            if formato == 'pdf':
                return self.exportar_pdf(data)
            elif formato == 'excel':
                return self.exportar_excel(data)
            else:
                return Response(data, status=status.HTTP_200_OK)
                
        except Exception as e:
            return Response(
                {'error': f'Error al generar reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def exportar_pdf(self, data):
        """Exporta el reporte a PDF con gráficas"""
        if not REPORTLAB_AVAILABLE:
            return Response({'error': 'ReportLab no disponible'}, status=500)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=40, rightMargin=40, topMargin=50, bottomMargin=50)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#4F46E5'), spaceAfter=6, alignment=TA_CENTER, fontName='Helvetica-Bold')
        elements.append(Paragraph('📦 Reporte de Inventario y Consumo', title_style))
        elements.append(Spacer(1, 6))
        
        fecha_style = ParagraphStyle('Fecha', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", fecha_style))
        elements.append(Spacer(1, 12))
        
        stats = data.get('estadisticas', {})
        
        # Tarjetas de estadísticas
        stats_data = [
            [
                Paragraph(f"<para align=center><b><font size=22 color='#4F46E5'>{stats.get('total_materias', 0)}</font></b><br/><font size=9 color='gray'>Total Materias</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=22 color='#EF4444'>{stats.get('stock_bajo', 0)}</font></b><br/><font size=9 color='gray'>Stock Bajo</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=22 color='#10B981'>{stats.get('stock_disponible', 0)}</font></b><br/><font size=9 color='gray'>Disponibles</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=20 color='#F59E0B'>{stats.get('consumo_total', 0):.1f}</font></b><br/><font size=9 color='gray'>Consumo Total</font></para>", styles['Normal'])
            ]
        ]
        stats_table = Table(stats_data, colWidths=[1.3*inch]*4)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#EEF2FF')),
            ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#FEE2E2')),
            ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#D1FAE5')),
            ('BACKGROUND', (3, 0), (3, 0), colors.HexColor('#FEF3C7')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 25))
        
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#4F46E5'), spaceAfter=10)
        
        # Gráfica de distribución por estado (Pie Chart)
        if stats.get('por_estado'):
            elements.append(Paragraph('📊 Distribución por Estado', subtitle_style))
            elements.append(Spacer(1, 10))
            
            drawing = Drawing(400, 180)
            pie = Pie()
            pie.x = 120
            pie.y = 20
            pie.width = 120
            pie.height = 120
            pie.data = list(stats['por_estado'].values())
            pie.labels = list(stats['por_estado'].keys())
            pie.slices.strokeWidth = 0.5
            
            colores = [colors.HexColor('#4F46E5'), colors.HexColor('#10B981'), colors.HexColor('#EF4444'), colors.HexColor('#F59E0B')]
            for i, color in enumerate(colores[:len(pie.data)]):
                pie.slices[i].fillColor = color
            
            drawing.add(pie)
            elements.append(drawing)
            elements.append(Spacer(1, 20))
        
        # Materiales críticos (Stock bajo)
        if stats.get('materiales_criticos'):
            elements.append(Paragraph('⚠️ Materiales en Stock Crítico', subtitle_style))
            elements.append(Spacer(1, 10))
            
            crit_data = [['Material', 'Stock Actual', 'Stock Mínimo', 'Unidad']]
            for m in stats['materiales_criticos']:
                crit_data.append([
                    m.get('nombre_materia_prima', '')[:25],
                    f"{m.get('stock_total', 0):.2f}",
                    f"{m.get('stock_minimo', 0):.2f}",
                    m.get('unidad_medida', '')
                ])
            
            crit_table = Table(crit_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch])
            crit_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EF4444')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#FEE2E2'), colors.white])
            ]))
            elements.append(crit_table)
            elements.append(Spacer(1, 20))
        
        # Stock de Materias Primas
        elements.append(Paragraph('📋 Stock de Materias Primas', subtitle_style))
        elements.append(Spacer(1, 10))
        
        table_data = [['Materia Prima', 'Stock Total', 'Unidad', 'Estado']]
        for item in data['stock_materias_primas'][:15]:  # Limitar a 15 para no saturar
            table_data.append([
                item['nombre_materia_prima'][:25],
                f"{item['stock_total']:.2f}",
                item['unidad_medida'],
                item['estado'][:12]
            ])
        
        table = Table(table_data, colWidths=[2.3*inch, 1*inch, 0.9*inch, 1.1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
        elements.append(PageBreak())
        
        # Consumo de Materiales
        elements.append(Paragraph('🔄 Consumo de Materiales', subtitle_style))
        elements.append(Spacer(1, 10))
        
        # Gráfica de barras de consumo
        consumo_data = data['consumo_materiales']['data'][:10]
        if consumo_data:
            drawing = Drawing(450, 180)
            bc = VerticalBarChart()
            bc.x = 40
            bc.y = 20
            bc.height = 140
            bc.width = 350
            bc.data = [[c.get('consumo_total', 0) for c in consumo_data]]
            bc.categoryAxis.categoryNames = [c.get('nombre_materia', '')[:10] for c in consumo_data]
            bc.bars[0].fillColor = colors.HexColor('#10B981')
            bc.valueAxis.valueMin = 0
            bc.categoryAxis.labels.boxAnchor = 'n'
            bc.categoryAxis.labels.angle = 30
            bc.categoryAxis.labels.fontSize = 7
            bc.valueAxis.labels.fontName = 'Helvetica'
            drawing.add(bc)
            elements.append(drawing)
            elements.append(Spacer(1, 15))
        
        table_data = [['Material', 'Consumo Total', 'Unidad']]
        for item in consumo_data:
            table_data.append([
                item['nombre_materia'][:30],
                f"{item['consumo_total']:.2f}",
                item['unidad_medida']
            ])
        
        table = Table(table_data, colWidths=[3*inch, 1.3*inch, 1.3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(f"TextilTech © {datetime.now().year} - Reporte generado automáticamente", footer_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Inventario_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response
    
    def exportar_excel(self, data):
        """Exporta el reporte a Excel con formato profesional"""
        if not OPENPYXL_AVAILABLE:
            return Response({'error': 'openpyxl no disponible'}, status=500)
        
        wb = Workbook()
        
        # === HOJA 1: RESUMEN ===
        ws_resumen = wb.active
        ws_resumen.title = "Resumen"
        
        # Título
        ws_resumen['A1'] = 'REPORTE DE INVENTARIO Y CONSUMO'
        ws_resumen['A1'].font = Font(bold=True, size=20, color="4F46E5")
        ws_resumen['A1'].alignment = Alignment(horizontal='center')
        ws_resumen.merge_cells('A1:E1')
        ws_resumen.row_dimensions[1].height = 30
        
        # Fecha
        ws_resumen['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws_resumen['A2'].font = Font(italic=True, size=10, color="666666")
        ws_resumen.merge_cells('A2:E2')
        
        # Estadísticas
        stats = data.get('estadisticas', {})
        row = 4
        
        kpis = [
            ('Total Materias Primas', stats.get('total_materias', 0), 'E8EAF6'),
            ('Stock Bajo', stats.get('stock_bajo', 0), 'FFCDD2'),
            ('Stock Disponible', stats.get('stock_disponible', 0), 'C8E6C9'),
            ('Consumo Total', f"{stats.get('consumo_total', 0):.2f}", 'FFF9C4')
        ]
        
        for label, value, color in kpis:
            ws_resumen[f'A{row}'] = label
            ws_resumen[f'A{row}'].font = Font(bold=True, size=11)
            ws_resumen[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws_resumen[f'B{row}'] = value
            ws_resumen[f'B{row}'].font = Font(size=11, bold=True, color="1976D2")
            ws_resumen[f'B{row}'].alignment = Alignment(horizontal='center')
            ws_resumen[f'B{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            row += 1
        
        # Materiales críticos
        if stats.get('materiales_criticos'):
            row += 2
            ws_resumen[f'A{row}'] = '⚠️ MATERIALES EN STOCK CRÍTICO'
            ws_resumen[f'A{row}'].font = Font(bold=True, size=12, color="EF4444")
            ws_resumen.merge_cells(f'A{row}:D{row}')
            row += 1
            
            headers = ['Material', 'Stock Actual', 'Stock Mínimo', 'Unidad']
            for col_num, header in enumerate(headers, 1):
                cell = ws_resumen.cell(row=row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            for mat in stats['materiales_criticos']:
                ws_resumen[f'A{row}'] = mat.get('nombre_materia_prima', '')[:30]
                ws_resumen[f'B{row}'] = mat.get('stock_total', 0)
                ws_resumen[f'C{row}'] = mat.get('stock_minimo', 0)
                ws_resumen[f'D{row}'] = mat.get('unidad_medida', '')
                for col in ['A', 'B', 'C', 'D']:
                    ws_resumen[f'{col}{row}'].fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                row += 1
        
        ws_resumen.column_dimensions['A'].width = 40
        ws_resumen.column_dimensions['B'].width = 15
        ws_resumen.column_dimensions['C'].width = 15
        ws_resumen.column_dimensions['D'].width = 12
        
        # === HOJA 2: STOCK MATERIAS PRIMAS ===
        ws_stock = wb.create_sheet("Stock Materias Primas")
        
        # Encabezados
        headers = ['Materia Prima', 'Stock Total', 'Stock Mínimo', 'Unidad', 'Estado']
        for col_num, header in enumerate(headers, 1):
            cell = ws_stock.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for row_num, item in enumerate(data['stock_materias_primas'], 2):
            ws_stock.cell(row=row_num, column=1, value=item['nombre_materia_prima'])
            ws_stock.cell(row=row_num, column=2, value=item['stock_total'])
            ws_stock.cell(row=row_num, column=3, value=item.get('stock_minimo', 0))
            ws_stock.cell(row=row_num, column=4, value=item['unidad_medida'])
            ws_stock.cell(row=row_num, column=5, value=item['estado'])
            
            # Formato numérico
            ws_stock.cell(row=row_num, column=2).number_format = '#,##0.00'
            ws_stock.cell(row=row_num, column=3).number_format = '#,##0.00'
            
            # Color según estado
            if item['stock_total'] < item.get('stock_minimo', 0):
                fill_color = 'FFCDD2'  # Rojo claro
            else:
                fill_color = 'F5F5F5' if row_num % 2 == 0 else 'FFFFFF'
            
            for col in range(1, 6):
                cell = ws_stock.cell(row=row_num, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
        
        ws_stock.column_dimensions['A'].width = 35
        ws_stock.column_dimensions['B'].width = 15
        ws_stock.column_dimensions['C'].width = 15
        ws_stock.column_dimensions['D'].width = 12
        ws_stock.column_dimensions['E'].width = 15
        
        # === HOJA 3: CONSUMO DE MATERIALES ===
        ws_consumo = wb.create_sheet("Consumo Materiales")
        
        # Encabezados
        headers = ['Material', 'Consumo Total', 'Unidad']
        for col_num, header in enumerate(headers, 1):
            cell = ws_consumo.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for row_num, item in enumerate(data['consumo_materiales']['data'], 2):
            ws_consumo.cell(row=row_num, column=1, value=item['nombre_materia'])
            ws_consumo.cell(row=row_num, column=2, value=item['consumo_total'])
            ws_consumo.cell(row=row_num, column=3, value=item['unidad_medida'])
            
            # Formato numérico
            ws_consumo.cell(row=row_num, column=2).number_format = '#,##0.00'
            
            # Color alternado
            fill_color = 'F5F5F5' if row_num % 2 == 0 else 'FFFFFF'
            for col in range(1, 4):
                cell = ws_consumo.cell(row=row_num, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
        
        ws_consumo.column_dimensions['A'].width = 40
        ws_consumo.column_dimensions['B'].width = 15
        ws_consumo.column_dimensions['C'].width = 12
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Inventario_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response


class ReporteClientesView(APIView):
    """
    Genera reporte de clientes
    GET /reportes/clientes/?formato=json|pdf|excel
    """
    def get(self, request):
        formato = request.query_params.get('formato', 'json')
        
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT 
                        id,
                        nombre_completo,
                        direccion,
                        telefono,
                        fecha_nacimiento,
                        estado
                    FROM clientes
                    ORDER BY nombre_completo
                """)
                columns = [col[0] for col in cursor.description]
                clientes = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Convertir fecha a string
            for cliente in clientes:
                if cliente.get('fecha_nacimiento'):
                    cliente['fecha_nacimiento'] = str(cliente['fecha_nacimiento'])
            
            # Estadísticas avanzadas
            total = len(clientes)
            activos = sum(1 for c in clientes if c.get('estado') == 'Activo')
            inactivos = total - activos
            
            # Agrupar por estado
            por_estado = {}
            for c in clientes:
                est = c.get('estado', 'Desconocido')
                if est not in por_estado:
                    por_estado[est] = 0
                por_estado[est] += 1
            
            data = {
                'fecha_generacion': datetime.now().isoformat(),
                'clientes': clientes,
                'estadisticas': {
                    'total_clientes': total,
                    'activos': activos,
                    'inactivos': inactivos,
                    'por_estado': por_estado
                }
            }
            
            if formato == 'pdf':
                return self.exportar_pdf(data)
            else:
                return Response(data)
        except Exception as e:
            return Response({'error': str(e)}, status=500)
    
    def exportar_pdf(self, data):
        """Exporta el reporte a PDF con gráficas"""
        if not REPORTLAB_AVAILABLE:
            return Response({'error': 'ReportLab no disponible'}, status=500)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=50, rightMargin=50)
        elements = []
        styles = getSampleStyleSheet()
        # Título
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#4F46E5'), spaceAfter=6, alignment=TA_CENTER, fontName='Helvetica-Bold')
        elements.append(Paragraph('👥 Reporte de Clientes', title_style))
        elements.append(Spacer(1, 6))
        fecha_style = ParagraphStyle('Fecha', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", fecha_style))
        elements.append(Spacer(1, 12))
        stats = data['estadisticas']
        # Tarjetas de estadísticas
        stats_data = [
            [
                Paragraph(f"<para align=center><b><font size=24 color='#4F46E5'>{stats['total_clientes']}</font></b><br/><font size=10 color='gray'>Total Clientes</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=24 color='#10B981'>{stats['activos']}</font></b><br/><font size=10 color='gray'>Activos</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=24 color='#EF4444'>{stats['inactivos']}</font></b><br/><font size=10 color='gray'>Inactivos</font></para>", styles['Normal'])
            ]
        ]
        stats_table = Table(stats_data, colWidths=[2*inch]*3)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#EEF2FF')),
            ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#D1FAE5')),
            ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#FEE2E2')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 15),
            ('RIGHTPADDING', (0, 0), (-1, -1), 15),
            ('TOPPADDING', (0, 0), (-1, -1), 15),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 30))
        
        # Gráfica Pie de distribución
        if stats.get('por_estado'):
            subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#4F46E5'), spaceAfter=10)
            elements.append(Paragraph('📊 Distribución por Estado', subtitle_style))
            elements.append(Spacer(1, 10))
            
            drawing = Drawing(400, 180)
            pie = Pie()
            pie.x = 130
            pie.y = 20
            pie.width = 120
            pie.height = 120
            pie.data = list(stats['por_estado'].values())
            pie.labels = list(stats['por_estado'].keys())
            pie.slices.strokeWidth = 0.5
            pie.slices[0].fillColor = colors.HexColor('#10B981')
            if len(pie.data) > 1:
                pie.slices[1].fillColor = colors.HexColor('#EF4444')
            drawing.add(pie)
            elements.append(drawing)
            elements.append(Spacer(1, 20))
        
        # Tabla de clientes
        elements.append(Paragraph('📋 Listado de Clientes', subtitle_style))
        elements.append(Spacer(1, 10))
        
        table_data = [['Nombre', 'Teléfono', 'Dirección', 'Estado']]
        for c in data['clientes'][:20]:  # Limitar a 20 para no saturar
            table_data.append([
                c.get('nombre_completo', '')[:25],
                c.get('telefono', '')[:15],
                c.get('direccion', '')[:30],
                c.get('estado', '')
            ])
        
        table = Table(table_data, colWidths=[2*inch, 1.2*inch, 2*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(f"TextilTech © {datetime.now().year} - Reporte generado automáticamente", footer_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Clientes_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response


class ReporteBitacoraView(APIView):
    """
    Genera reporte de bitácora
    GET /reportes/bitacora/?fecha_inicio=YYYY-MM-DD&fecha_fin=YYYY-MM-DD&formato=json|pdf|excel
    """
    def get(self, request):
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        formato = request.query_params.get('formato', 'json')
        
        try:
            with connection.cursor() as cursor:
                query = """
                    SELECT 
                        username,
                        ip,
                        fecha_hora,
                        accion,
                        descripcion
                    FROM bitacora
                    WHERE 1=1
                """
                params = []
                if fecha_inicio:
                    query += " AND fecha_hora >= %s"
                    params.append(fecha_inicio)
                if fecha_fin:
                    query += " AND fecha_hora <= %s"
                    params.append(fecha_fin + ' 23:59:59')
                
                query += " ORDER BY fecha_hora DESC"
                
                cursor.execute(query, params)
                columns = [col[0] for col in cursor.description]
                actividades = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Convertir fecha a string
            for act in actividades:
                if act.get('fecha_hora'):
                    act['fecha_hora'] = str(act['fecha_hora'])
            
            # Estadísticas avanzadas
            usuarios_activos = len(set(a['username'] for a in actividades))
            
            # Agrupar por usuario
            por_usuario = {}
            for a in actividades:
                usr = a.get('username', 'Desconocido')
                if usr not in por_usuario:
                    por_usuario[usr] = 0
                por_usuario[usr] += 1
            
            # Agrupar por acción
            por_accion = {}
            for a in actividades:
                acc = a.get('accion', 'Desconocida')
                if acc not in por_accion:
                    por_accion[acc] = 0
                por_accion[acc] += 1
            
            data = {
                'fecha_generacion': datetime.now().isoformat(),
                'actividades': actividades,
                'estadisticas': {
                    'total_actividades': len(actividades),
                    'usuarios_activos': usuarios_activos,
                    'por_usuario': por_usuario,
                    'por_accion': por_accion
                },
                'filtros': {
                    'fecha_inicio': fecha_inicio,
                    'fecha_fin': fecha_fin
                }
            }
            
            if formato == 'pdf':
                return self.exportar_pdf(data)
            elif formato == 'excel':
                return self.exportar_excel(data)
            else:
                return Response(data)
        except Exception as e:
            return Response({'error': str(e)}, status=500)
    
    def exportar_pdf(self, data):
        """Exporta el reporte a PDF con gráficas"""
        if not REPORTLAB_AVAILABLE:
            return Response({'error': 'ReportLab no disponible'}, status=500)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=50, rightMargin=50)
        elements = []
        styles = getSampleStyleSheet()
        # Título
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#4F46E5'), spaceAfter=6, alignment=TA_CENTER, fontName='Helvetica-Bold')
        elements.append(Paragraph('📝 Reporte de Bitácora', title_style))
        elements.append(Spacer(1, 6))
        fecha_style = ParagraphStyle('Fecha', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", fecha_style))
        elements.append(Spacer(1, 12))
        stats = data['estadisticas']
        
        # Tarjetas de estadísticas
        stats_data = [
            [
                Paragraph(f"<para align=center><b><font size=24 color='#4F46E5'>{stats['total_actividades']}</font></b><br/><font size=10 color='gray'>Total Actividades</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=24 color='#10B981'>{stats['usuarios_activos']}</font></b><br/><font size=10 color='gray'>Usuarios Activos</font></para>", styles['Normal'])
            ]
        ]
        stats_table = Table(stats_data, colWidths=[3*inch]*2)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#EEF2FF')),
            ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#D1FAE5')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 15),
            ('RIGHTPADDING', (0, 0), (-1, -1), 15),
            ('TOPPADDING', (0, 0), (-1, -1), 15),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 30))
        
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#4F46E5'), spaceAfter=10)
        
        # Gráfica de barras por acción
        if stats.get('por_accion'):
            elements.append(Paragraph('📊 Actividades por Acción', subtitle_style))
            elements.append(Spacer(1, 10))
            
            drawing = Drawing(450, 180)
            bc = VerticalBarChart()
            bc.x = 40
            bc.y = 20
            bc.height = 140
            bc.width = 350
            sorted_acciones = sorted(stats['por_accion'].items(), key=lambda x: x[1], reverse=True)[:8]
            bc.data = [[v for k, v in sorted_acciones]]
            bc.categoryAxis.categoryNames = [k[:12] for k, v in sorted_acciones]
            bc.bars[0].fillColor = colors.HexColor('#4F46E5')
            bc.valueAxis.valueMin = 0
            bc.categoryAxis.labels.boxAnchor = 'n'
            bc.categoryAxis.labels.angle = 30
            bc.categoryAxis.labels.fontSize = 8
            drawing.add(bc)
            elements.append(drawing)
            elements.append(Spacer(1, 20))
        
        # Tabla de actividades
        elements.append(Paragraph('📋 Registro de Actividades', subtitle_style))
        elements.append(Spacer(1, 10))
        
        table_data = [['Usuario', 'IP', 'Fecha/Hora', 'Acción']]
        for act in data['actividades'][:30]:  # Limitar a 30 registros
            table_data.append([
                act.get('username', '')[:15],
                act.get('ip', '')[:15],
                act.get('fecha_hora', '')[:19],
                act.get('accion', '')[:20]
            ])
        
        table = Table(table_data, colWidths=[1.3*inch, 1.2*inch, 1.5*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
        ]))
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(f"TextilTech © {datetime.now().year} - Reporte generado automáticamente", footer_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Bitacora_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response
    
    def exportar_excel(self, data):
        """Exporta el reporte de bitácora a Excel con formato profesional"""
        if not OPENPYXL_AVAILABLE:
            return Response({'error': 'openpyxl no disponible'}, status=500)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bitácora"
        
        # Título
        ws['A1'] = 'REPORTE DE BITÁCORA'
        ws['A1'].font = Font(bold=True, size=20, color="4F46E5")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 30
        
        # Fecha
        ws['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, size=10, color="666666")
        ws.merge_cells('A2:E2')
        
        # Estadísticas
        stats = data['estadisticas']
        ws['A4'] = 'Total Actividades'
        ws['B4'] = stats['total_actividades']
        ws['C4'] = 'Usuarios Activos'
        ws['D4'] = stats['usuarios_activos']
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}4'].font = Font(bold=True)
            ws[f'{col}4'].fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
        
        # Encabezados de tabla
        headers = ['Usuario', 'IP', 'Fecha/Hora', 'Acción', 'Descripción']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for row_num, act in enumerate(data['actividades'], 7):
            ws.cell(row=row_num, column=1, value=act.get('username', ''))
            ws.cell(row=row_num, column=2, value=act.get('ip', ''))
            ws.cell(row=row_num, column=3, value=act.get('fecha_hora', '')[:19] if act.get('fecha_hora') else '')
            ws.cell(row=row_num, column=4, value=act.get('accion', ''))
            ws.cell(row=row_num, column=5, value=act.get('descripcion', '')[:50])
            
            # Color alternado
            fill_color = 'F5F5F5' if row_num % 2 == 0 else 'FFFFFF'
            for col in range(1, 6):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 50
        
        # === GRÁFICA DE BARRAS - ACCIONES ===
        if stats.get('por_accion'):
            chart_sheet = wb.create_sheet("📊 Gráficas")
            
            # Datos para la gráfica
            chart_sheet['A1'] = 'Acción'
            chart_sheet['B1'] = 'Cantidad'
            row = 2
            for accion, cantidad in stats['por_accion'].items():
                chart_sheet[f'A{row}'] = accion
                chart_sheet[f'B{row}'] = cantidad
                row += 1
            
            # Crear gráfica de barras
            chart = BarChart()
            chart.title = "Distribución de Acciones en la Bitácora"
            chart.style = 10
            chart.y_axis.title = 'Cantidad'
            chart.x_axis.title = 'Tipo de Acción'
            
            data = Reference(chart_sheet, min_col=2, min_row=1, max_row=row-1)
            cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=row-1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 20
            
            chart_sheet.add_chart(chart, "D2")
            
            # Gráfica de pastel
            pie = PieChart()
            pie.title = "Distribución Porcentual de Acciones"
            labels = Reference(chart_sheet, min_col=1, min_row=2, max_row=row-1)
            data = Reference(chart_sheet, min_col=2, min_row=1, max_row=row-1)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.height = 10
            pie.width = 12
            
            chart_sheet.add_chart(pie, "D22")
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Bitacora_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response


class ReportePersonalView(APIView):
    """
    Genera reporte de personal
    GET /reportes/personal/?formato=json|pdf|excel
    """
    def get(self, request):
        formato = request.query_params.get('formato', 'json')
        
        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT 
                        id,
                        nombre_completo,
                        telefono,
                        '' as ci,
                        rol,
                        direccion,
                        fecha_nacimiento as fecha_contratacion,
                        estado
                    FROM personal
                    ORDER BY nombre_completo
                """)
                columns = [col[0] for col in cursor.description]
                empleados = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Convertir fecha a string
            for emp in empleados:
                if emp.get('fecha_contratacion'):
                    emp['fecha_contratacion'] = str(emp['fecha_contratacion'])
            
            # Estadísticas avanzadas
            total = len(empleados)
            activos = sum(1 for e in empleados if e.get('estado') == 'Activo')
            inactivos = total - activos
            
            # Agrupar por rol
            por_rol = {}
            for e in empleados:
                rol = e.get('rol', 'Sin rol')
                if rol not in por_rol:
                    por_rol[rol] = 0
                por_rol[rol] += 1
            
            data = {
                'fecha_generacion': datetime.now().isoformat(),
                'empleados': empleados,
                'estadisticas': {
                    'total_empleados': total,
                    'activos': activos,
                    'inactivos': inactivos,
                    'por_rol': por_rol
                }
            }
            
            if formato == 'pdf':
                return self.exportar_pdf(data)
            elif formato == 'excel':
                return self.exportar_excel(data)
            else:
                return Response(data)
        except Exception as e:
            return Response({'error': str(e)}, status=500)
    
    def exportar_pdf(self, data):
        """Exporta el reporte a PDF con gráficas"""
        if not REPORTLAB_AVAILABLE:
            return Response({'error': 'ReportLab no disponible'}, status=500)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=50, rightMargin=50)
        elements = []
        styles = getSampleStyleSheet()
        # Título
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#4F46E5'), spaceAfter=6, alignment=TA_CENTER, fontName='Helvetica-Bold')
        elements.append(Paragraph('👤 Reporte de Personal', title_style))
        elements.append(Spacer(1, 6))
        fecha_style = ParagraphStyle('Fecha', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", fecha_style))
        elements.append(Spacer(1, 12))
        stats = data['estadisticas']
        
        # Tarjetas de estadísticas
        stats_data = [
            [
                Paragraph(f"<para align=center><b><font size=24 color='#4F46E5'>{stats['total_empleados']}</font></b><br/><font size=10 color='gray'>Total Empleados</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=24 color='#10B981'>{stats['activos']}</font></b><br/><font size=10 color='gray'>Activos</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=24 color='#EF4444'>{stats['inactivos']}</font></b><br/><font size=10 color='gray'>Inactivos</font></para>", styles['Normal'])
            ]
        ]
        stats_table = Table(stats_data, colWidths=[2*inch]*3)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#EEF2FF')),
            ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#D1FAE5')),
            ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#FEE2E2')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 15),
            ('RIGHTPADDING', (0, 0), (-1, -1), 15),
            ('TOPPADDING', (0, 0), (-1, -1), 15),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 30))
        
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#4F46E5'), spaceAfter=10)
        
        # Gráfica por rol
        if stats.get('por_rol'):
            elements.append(Paragraph('📊 Distribución por Rol', subtitle_style))
            elements.append(Spacer(1, 10))
            
            drawing = Drawing(400, 180)
            bc = VerticalBarChart()
            bc.x = 50
            bc.y = 20
            bc.height = 140
            bc.width = 300
            bc.data = [list(stats['por_rol'].values())]
            bc.categoryAxis.categoryNames = list(stats['por_rol'].keys())
            bc.bars[0].fillColor = colors.HexColor('#10B981')
            bc.valueAxis.valueMin = 0
            bc.categoryAxis.labels.boxAnchor = 'n'
            bc.categoryAxis.labels.angle = 30
            bc.categoryAxis.labels.fontSize = 8
            drawing.add(bc)
            elements.append(drawing)
            elements.append(Spacer(1, 20))
        
        # Tabla de empleados
        elements.append(Paragraph('📋 Listado de Personal', subtitle_style))
        elements.append(Spacer(1, 10))
        
        table_data = [['Nombre', 'Teléfono', 'Rol', 'Estado']]
        for emp in data['empleados'][:20]:
            table_data.append([
                emp.get('nombre_completo', '')[:25],
                emp.get('telefono', '')[:15],
                emp.get('rol', '')[:15],
                emp.get('estado', '')
            ])
        
        table = Table(table_data, colWidths=[2.5*inch, 1.5*inch, 1.2*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(f"TextilTech © {datetime.now().year} - Reporte generado automáticamente", footer_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Personal_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response
    
    def exportar_excel(self, data):
        """Exporta el reporte de personal a Excel con formato profesional"""
        if not OPENPYXL_AVAILABLE:
            return Response({'error': 'openpyxl no disponible'}, status=500)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Personal"
        
        # Título
        ws['A1'] = 'REPORTE DE PERSONAL'
        ws['A1'].font = Font(bold=True, size=20, color="4F46E5")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 30
        
        # Fecha
        ws['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, size=10, color="666666")
        ws.merge_cells('A2:E2')
        
        # Estadísticas
        stats = data['estadisticas']
        ws['A4'] = 'Total Empleados'
        ws['B4'] = stats['total_empleados']
        ws['C4'] = 'Activos'
        ws['D4'] = stats['activos']
        ws['E4'] = 'Inactivos'
        ws['F4'] = stats['inactivos']
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}4'].font = Font(bold=True)
            ws[f'{col}4'].fill = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
        
        # Encabezados de tabla
        headers = ['Nombre', 'Teléfono', 'Rol', 'Dirección', 'Estado']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for row_num, emp in enumerate(data['empleados'], 7):
            ws.cell(row=row_num, column=1, value=emp.get('nombre_completo', ''))
            ws.cell(row=row_num, column=2, value=emp.get('telefono', ''))
            ws.cell(row=row_num, column=3, value=emp.get('rol', ''))
            ws.cell(row=row_num, column=4, value=emp.get('direccion', ''))
            ws.cell(row=row_num, column=5, value=emp.get('estado', ''))
            
            # Color alternado
            fill_color = 'F5F5F5' if row_num % 2 == 0 else 'FFFFFF'
            for col in range(1, 6):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 12
        
        # === GRÁFICA DE BARRAS - DISTRIBUCIÓN POR ROL ===
        if stats.get('por_rol'):
            chart_sheet = wb.create_sheet("📊 Gráficas")
            
            # Datos para la gráfica
            chart_sheet['A1'] = 'Rol'
            chart_sheet['B1'] = 'Cantidad'
            row = 2
            for rol, cantidad in stats['por_rol'].items():
                chart_sheet[f'A{row}'] = rol
                chart_sheet[f'B{row}'] = cantidad
                row += 1
            
            # Crear gráfica de barras
            chart = BarChart()
            chart.title = "Distribución de Personal por Rol"
            chart.style = 11
            chart.y_axis.title = 'Cantidad de Empleados'
            chart.x_axis.title = 'Rol'
            
            data = Reference(chart_sheet, min_col=2, min_row=1, max_row=row-1)
            cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=row-1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 20
            
            chart_sheet.add_chart(chart, "D2")
            
            # Gráfica de pastel - Activos vs Inactivos
            chart_sheet['E1'] = 'Estado'
            chart_sheet['F1'] = 'Cantidad'
            chart_sheet['E2'] = 'Activos'
            chart_sheet['F2'] = stats['activos']
            chart_sheet['E3'] = 'Inactivos'
            chart_sheet['F3'] = stats['inactivos']
            
            pie = PieChart()
            pie.title = "Empleados Activos vs Inactivos"
            labels = Reference(chart_sheet, min_col=5, min_row=2, max_row=3)
            data = Reference(chart_sheet, min_col=6, min_row=1, max_row=3)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.height = 10
            pie.width = 12
            
            chart_sheet.add_chart(pie, "D22")
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Personal_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response


class ReportePedidosView(APIView):
    """
    Genera reporte de pedidos
    GET /reportes/pedidos/?fecha_inicio=YYYY-MM-DD&fecha_fin=YYYY-MM-DD&formato=json|pdf|excel
    """
    def get(self, request):
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        formato = request.query_params.get('formato', 'json')
        
        try:
            with connection.cursor() as cursor:
                query = """
                    SELECT 
                        p.id_pedido as id,
                        p.cod_pedido as codigo_pedido,
                        c.nombre_completo as cliente,
                        p.fecha_pedido,
                        p.fecha_entrega_prometida as fecha_entrega,
                        p.estado,
                        p.total
                    FROM pedidos p
                    LEFT JOIN clientes c ON p.id_cliente = c.id
                    WHERE 1=1
                """
                params = []
                if fecha_inicio:
                    query += " AND p.fecha_pedido >= %s"
                    params.append(fecha_inicio)
                if fecha_fin:
                    query += " AND p.fecha_pedido <= %s"
                    params.append(fecha_fin)
                
                query += " ORDER BY p.fecha_pedido DESC"
                
                cursor.execute(query, params)
                columns = [col[0] for col in cursor.description]
                pedidos = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Convertir fechas y decimales
            for pedido in pedidos:
                if pedido.get('fecha_pedido'):
                    pedido['fecha_pedido'] = str(pedido['fecha_pedido'])
                if pedido.get('fecha_entrega'):
                    pedido['fecha_entrega'] = str(pedido['fecha_entrega'])
                if pedido.get('total'):
                    pedido['total'] = float(pedido['total'])
            
            # Estadísticas avanzadas
            total_pedidos = len(pedidos)
            pendientes = sum(1 for p in pedidos if p.get('estado') == 'Pendiente')
            completados = sum(1 for p in pedidos if p.get('estado') == 'Completado')
            cancelados = sum(1 for p in pedidos if p.get('estado') == 'Cancelado')
            monto_total = sum(p.get('total', 0) for p in pedidos)
            
            # Agrupar por estado
            por_estado = {}
            for p in pedidos:
                est = p.get('estado', 'Desconocido')
                if est not in por_estado:
                    por_estado[est] = 0
                por_estado[est] += 1
            
            # Promedio por pedido
            promedio = (monto_total / total_pedidos) if total_pedidos > 0 else 0
            
            data = {
                'fecha_generacion': datetime.now().isoformat(),
                'pedidos': pedidos,
                'estadisticas': {
                    'total_pedidos': total_pedidos,
                    'pendientes': pendientes,
                    'completados': completados,
                    'cancelados': cancelados,
                    'monto_total': monto_total,
                    'promedio': promedio,
                    'por_estado': por_estado
                },
                'filtros': {
                    'fecha_inicio': fecha_inicio,
                    'fecha_fin': fecha_fin
                }
            }
            
            if formato == 'pdf':
                return self.exportar_pdf(data)
            elif formato == 'excel':
                return self.exportar_excel(data)
            else:
                return Response(data)
        except Exception as e:
            return Response({'error': str(e)}, status=500)
    
    def exportar_pdf(self, data):
        """Exporta el reporte a PDF con gráficas"""
        if not REPORTLAB_AVAILABLE:
            return Response({'error': 'ReportLab no disponible'}, status=500)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=50, rightMargin=50)
        elements = []
        styles = getSampleStyleSheet()
        # Título
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#4F46E5'), spaceAfter=6, alignment=TA_CENTER, fontName='Helvetica-Bold')
        elements.append(Paragraph('🛒 Reporte de Pedidos', title_style))
        elements.append(Spacer(1, 6))
        fecha_style = ParagraphStyle('Fecha', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", fecha_style))
        elements.append(Spacer(1, 12))
        stats = data['estadisticas']
        
        # Tarjetas de estadísticas (4 KPIs)
        stats_data = [
            [
                Paragraph(f"<para align=center><b><font size=16 color='#4F46E5'>{stats['total_pedidos']}</font></b><br/><font size=8 color='gray'>Total Pedidos</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=16 color='#F59E0B'>{stats['pendientes']}</font></b><br/><font size=8 color='gray'>Pendientes</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=16 color='#10B981'>{stats['completados']}</font></b><br/><font size=8 color='gray'>Completados</font></para>", styles['Normal']),
                Paragraph(f"<para align=center><b><font size=14 color='#6366F1'>Bs.{stats['monto_total']:.2f}</font></b><br/><font size=8 color='gray'>Monto Total</font></para>", styles['Normal'])
            ]
        ]
        stats_table = Table(stats_data, colWidths=[1.3*inch]*4)
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#EEF2FF')),
            ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#FEF3C7')),
            ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#D1FAE5')),
            ('BACKGROUND', (3, 0), (3, 0), colors.HexColor('#E0E7FF')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(stats_table)
        elements.append(Spacer(1, 18))
        
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#4F46E5'), spaceAfter=10)
        
        # Gráfica de barras por estado
        if stats.get('por_estado'):
            elements.append(Paragraph('📊 Distribución por Estado', subtitle_style))
            elements.append(Spacer(1, 10))
            
            drawing = Drawing(400, 180)
            bc = VerticalBarChart()
            bc.x = 50
            bc.y = 20
            bc.height = 140
            bc.width = 300
            bc.data = [list(stats['por_estado'].values())]
            bc.categoryAxis.categoryNames = list(stats['por_estado'].keys())
            bc.bars[0].fillColor = colors.HexColor('#10B981')
            bc.valueAxis.valueMin = 0
            bc.categoryAxis.labels.boxAnchor = 'n'
            bc.categoryAxis.labels.angle = 30
            bc.categoryAxis.labels.fontSize = 9
            drawing.add(bc)
            elements.append(drawing)
            elements.append(Spacer(1, 20))
        
        # Tabla de distribución con porcentajes
        elements.append(Paragraph('📈 Estadísticas de Pedidos', subtitle_style))
        elements.append(Spacer(1, 10))
        
        dist_data = [['Estado', 'Cantidad', 'Porcentaje']]
        for estado, cant in stats['por_estado'].items():
            porcentaje = (cant / stats['total_pedidos'] * 100) if stats['total_pedidos'] > 0 else 0
            dist_data.append([estado, str(cant), f"{porcentaje:.1f}%"])
        
        dist_table = Table(dist_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        dist_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')])
        ]))
        elements.append(dist_table)
        elements.append(Spacer(1, 25))
        
        # Tabla detallada de pedidos
        elements.append(Paragraph('📋 Detalle de Pedidos', subtitle_style))
        elements.append(Spacer(1, 10))
        
        table_data = [['Código', 'Cliente', 'F. Pedido', 'Estado', 'Total']]
        for p in data['pedidos'][:20]:  # Limitar a 20
            table_data.append([
                p.get('codigo_pedido', '')[:12],
                p.get('cliente', '')[:20],
                p.get('fecha_pedido', '')[:10],
                p.get('estado', '')[:12],
                f"Bs.{p.get('total', 0):.2f}"
            ])
        
        table = Table(table_data, colWidths=[1.2*inch, 2*inch, 1*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph(f"TextilTech © {datetime.now().year} - Reporte generado automáticamente", footer_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Pedidos_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response
    
    def exportar_excel(self, data):
        """Exporta el reporte de pedidos a Excel con formato profesional"""
        if not OPENPYXL_AVAILABLE:
            return Response({'error': 'openpyxl no disponible'}, status=500)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Pedidos"
        
        # Título
        ws['A1'] = 'REPORTE DE PEDIDOS'
        ws['A1'].font = Font(bold=True, size=20, color="4F46E5")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        # Fecha
        ws['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].font = Font(italic=True, size=10, color="666666")
        ws.merge_cells('A2:F2')
        
        # Estadísticas
        stats = data['estadisticas']
        row = 4
        kpis = [
            ('Total Pedidos', stats['total_pedidos'], 'EEF2FF'),
            ('Pendientes', stats['pendientes'], 'FEF3C7'),
            ('Completados', stats['completados'], 'D1FAE5'),
            ('Cancelados', stats['cancelados'], 'FEE2E2'),
            ('Monto Total (Bs.)', f"{stats['monto_total']:.2f}", 'E0E7FF'),
            ('Promedio (Bs.)', f"{stats['promedio']:.2f}", 'F3E8FF')
        ]
        
        col = 1
        for label, value, color in kpis:
            ws.cell(row=row, column=col, value=label).font = Font(bold=True, size=10)
            ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=row+1, column=col, value=value).font = Font(bold=True, size=14, color="4F46E5")
            ws.cell(row=row+1, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=row+1, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            col += 1
        
        # Encabezados de tabla
        headers = ['Código', 'Cliente', 'F. Pedido', 'F. Entrega', 'Estado', 'Total (Bs.)']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for row_num, pedido in enumerate(data['pedidos'], 8):
            ws.cell(row=row_num, column=1, value=pedido.get('codigo_pedido', ''))
            ws.cell(row=row_num, column=2, value=pedido.get('cliente', ''))
            ws.cell(row=row_num, column=3, value=pedido.get('fecha_pedido', '')[:10] if pedido.get('fecha_pedido') else '')
            ws.cell(row=row_num, column=4, value=pedido.get('fecha_entrega', '')[:10] if pedido.get('fecha_entrega') else '')
            ws.cell(row=row_num, column=5, value=pedido.get('estado', ''))
            ws.cell(row=row_num, column=6, value=pedido.get('total', 0))
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'
            
            # Color alternado
            fill_color = 'F5F5F5' if row_num % 2 == 0 else 'FFFFFF'
            for col in range(1, 7):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        
        # === GRÁFICA DE BARRAS - PEDIDOS POR ESTADO ===
        if stats.get('por_estado'):
            chart_sheet = wb.create_sheet("📊 Gráficas")
            
            # Datos para la gráfica
            chart_sheet['A1'] = 'Estado'
            chart_sheet['B1'] = 'Cantidad'
            chart_sheet['C1'] = 'Porcentaje'
            row = 2
            for estado, cantidad in stats['por_estado'].items():
                chart_sheet[f'A{row}'] = estado
                chart_sheet[f'B{row}'] = cantidad
                porcentaje = (cantidad / stats['total_pedidos'] * 100) if stats['total_pedidos'] > 0 else 0
                chart_sheet[f'C{row}'] = f"{porcentaje:.1f}%"
                row += 1
            
            # Crear gráfica de barras
            chart = BarChart()
            chart.title = "Distribución de Pedidos por Estado"
            chart.style = 12
            chart.y_axis.title = 'Cantidad de Pedidos'
            chart.x_axis.title = 'Estado'
            
            data = Reference(chart_sheet, min_col=2, min_row=1, max_row=row-1)
            cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=row-1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 20
            
            chart_sheet.add_chart(chart, "E2")
            
            # Gráfica de pastel
            pie = PieChart()
            pie.title = "Distribución Porcentual de Pedidos"
            labels = Reference(chart_sheet, min_col=1, min_row=2, max_row=row-1)
            data = Reference(chart_sheet, min_col=2, min_row=1, max_row=row-1)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.height = 10
            pie.width = 12
            
            chart_sheet.add_chart(pie, "E22")
            
            # KPIs visuales en la hoja de gráficas
            chart_sheet['A10'] = 'INDICADORES CLAVE'
            chart_sheet['A10'].font = Font(bold=True, size=14, color="4F46E5")
            
            chart_sheet['A11'] = 'Total de Pedidos:'
            chart_sheet['B11'] = stats['total_pedidos']
            chart_sheet['A12'] = 'Monto Total:'
            chart_sheet['B12'] = f"Bs. {stats['monto_total']:.2f}"
            chart_sheet['A13'] = 'Promedio por Pedido:'
            chart_sheet['B13'] = f"Bs. {stats['promedio']:.2f}"
            
            for cell in ['B11', 'B12', 'B13']:
                chart_sheet[cell].font = Font(bold=True, size=12, color="10B981")
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_Pedidos_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response
